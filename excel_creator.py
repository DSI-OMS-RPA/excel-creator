import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.drawing.image import Image
import csv
import json
import os
import logging

class ExcelCreator:
    """
    A class to create and manipulate Excel workbooks using openpyxl.
    This class provides various utilities like adding headers, data rows, formatting,
    conditional formatting, creating charts, data validation, and more.

    Attributes:
    - file_name (str): Name of the Excel file to create or append to.
    - include_header (bool): Whether to include headers when writing data.
    - append (bool): If True, append to an existing file, else create a new one.
    - header_font (Font): Custom font for headers (default is bold white).
    - header_bg_color (str): Background color for headers in hex format (default is light blue).
    """

    def __init__(self, file_name, include_header=True, append=False, header_font=None, header_bg_color=None):
        """
        Initialize an ExcelCreator object with optional header font and background color customization.

        Parameters:
        - file_name (str): Name of the Excel file to create or append to.
        - include_header (bool): Whether to include headers when writing data. Default is True.
        - append (bool): If True, append to an existing file. If False, create a new file. Default is False.
        - header_font (Font, optional): Custom font for headers. Default is bold white text.
        - header_bg_color (str, optional): Background color for headers in hex. Default is light blue.
        """
        self.file_name = file_name
        self.include_header = include_header
        self.append = append

        # Set default header font if none is provided
        self.header_font = header_font if header_font else Font(bold=True, color="FFFFFF")
        self.header_bg_color = header_bg_color if header_bg_color else "4F81BD"  # Default light blue

        # Load workbook if appending; otherwise, create a new workbook
        if self.append and os.path.exists(self.file_name):
            self.workbook = openpyxl.load_workbook(self.file_name)
        else:
            self.workbook = openpyxl.Workbook()

        # Activate the first sheet by default
        self.sheet = self.workbook.active

        # Set up logging
        logging.basicConfig(filename='excel_creator.log', level=logging.INFO)
        self.logger = logging.getLogger(__name__)

    def set_sheet_name(self, name):
        """
        Set the name of the active sheet.

        Parameters:
        - name (str): The new name for the sheet.
        """
        self.sheet.title = name

    def create_sheet(self, sheet_name, switch_to=False):
        """
        Create a new sheet in the workbook.

        Parameters:
        - sheet_name (str): The name of the new sheet.
        - switch_to (bool): If True, switch to the newly created sheet.
        """
        new_sheet = self.workbook.create_sheet(title=sheet_name)
        if switch_to:
            self.sheet = new_sheet

    def add_headers(self, headers, start_row=1):
        """
        Add headers to the specified row of the sheet.

        Parameters:
        - headers (list): List of column headers to add.
        - start_row (int): The row number to place the headers (default is row 1).
        """
        if self.include_header:
            for col_num, header in enumerate(headers, 1):
                cell = self.sheet.cell(row=start_row, column=col_num)
                cell.value = header
                # Apply custom font and background color
                cell.font = self.header_font
                cell.fill = PatternFill(start_color=self.header_bg_color, end_color=self.header_bg_color, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    def add_row(self, data, font=None, start_row=None):
        """
        Add a single row of data to the sheet.

        Parameters:
        - data (list): List of values for each column in the row.
        - font (Font, optional): Font style for the row data. If None, no custom font is applied.
        - start_row (int, optional): Row number to place the data. If None, data is appended.
        """
        if start_row:
            for col_num, value in enumerate(data, 1):
                cell = self.sheet.cell(row=start_row, column=col_num)
                cell.value = value
                if font:
                    cell.font = font
        else:
            self.sheet.append(data)
            if font:
                row_num = self.sheet.max_row
                for col_num, _ in enumerate(data, 1):
                    cell = self.sheet.cell(row=row_num, column=col_num)
                    cell.font = font

    def auto_number_rows(self, start_row=2, column=1):
        """
        Automatically number rows in a specific column, starting from a given row.

        Parameters:
        - start_row (int): The row to start numbering from (default is row 2).
        - column (int): The column to place the row numbers (default is column 1).
        """
        for row_num in range(start_row, self.sheet.max_row + 1):
            self.sheet.cell(row=row_num, column=column).value = row_num - start_row + 1

    def set_column_widths(self, auto_size=False, widths=None):
        """
        Set the width of each column, either automatically or using custom widths.

        Parameters:
        - auto_size (bool): If True, auto-sizes columns based on the longest entry.
        - widths (list): List of integers representing custom column widths. If None, auto-sizing is used.
        """
        if auto_size:
            for col in self.sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                self.sheet.column_dimensions[column].width = adjusted_width
        elif widths:
            for i, width in enumerate(widths, 1):
                column_letter = get_column_letter(i)
                self.sheet.column_dimensions[column_letter].width = width

    def apply_conditional_formatting(self, cell_range, min_color="FF0000", max_color="00FF00"):
        """
        Apply color scale conditional formatting to a range of cells.

        Parameters:
        - cell_range (str): The cell range to apply the formatting (e.g., "A1:A10").
        - min_color (str): The color for the minimum value in the scale (default is red).
        - max_color (str): The color for the maximum value in the scale (default is green).
        """
        rule = ColorScaleRule(start_type="min", start_color=min_color, end_type="max", end_color=max_color)
        self.sheet.conditional_formatting.add(cell_range, rule)

    def apply_custom_conditional_formatting(self, cell_range, operator, formula, font=None, bg_color=None):
        """
        Apply custom conditional formatting to a range of cells based on a formula.

        Parameters:
        - cell_range (str): The cell range to apply the formatting (e.g., "A1:A10").
        - operator (str): The operator for the condition (e.g., "greaterThan").
        - formula (str): The formula to apply (e.g., "5").
        - font (Font, optional): The font to apply if the condition is met.
        - bg_color (str, optional): Background color to apply if the condition is met.
        """
        rule = FormulaRule(formula=[formula], font=font, fill=PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid"))
        self.sheet.conditional_formatting.add(cell_range, rule)

    def add_data_validation(self, cell_range, validation_type="list", formula1=None):
        """
        Add data validation (e.g., a drop-down list) to a specified range of cells.

        Parameters:
        - cell_range (str): The cell range to apply data validation (e.g., "A1:A10").
        - validation_type (str): The type of validation (default is "list").
        - formula1 (str): Comma-separated list of valid options for the drop-down.
        """
        dv = DataValidation(type=validation_type, formula1=formula1, allow_blank=True)
        self.sheet.add_data_validation(dv)
        for row in self.sheet[cell_range]:
            for cell in row:
                dv.add(cell)

    def merge_cells(self, cell_range):
        """
        Merge a range of cells.

        Parameters:
        - cell_range (str): The cell range to merge (e.g., "A1:B1").
        """
        self.sheet.merge_cells(cell_range)

    def freeze_panes(self, cell):
        """
        Freeze panes at a specific cell (e.g., freeze row and column at "A2").

        Parameters:
        - cell (str): The cell to freeze panes at (e.g., "A2").
        """
        self.sheet.freeze_panes = self.sheet[cell]

    def create_chart(self, min_col, min_row, max_col, max_row, chart_type="bar", title=None, x_axis_title=None, y_axis_title=None, position="E15", include_legend=True, show_data_labels=False):
        """
        Create a chart based on the specified data range.

        Parameters:
        - min_col (int): Starting column of the data.
        - min_row (int): Starting row of the data.
        - max_col (int): Ending column of the data.
        - max_row (int): Ending row of the data.
        - chart_type (str): Type of chart to create ("bar", "line", "pie", "scatter"). Default is "bar".
        - title (str, optional): Title of the chart.
        - x_axis_title (str, optional): Title of the X-axis.
        - y_axis_title (str, optional): Title of the Y-axis.
        - position (str, optional): The cell location where the chart should be placed (default is "E15").
        - include_legend (bool): Whether to include a legend. Default is True.
        - show_data_labels (bool): Whether to show data labels on the chart. Default is False.
        """
        try:
            # Select the data for the chart
            data = Reference(self.sheet, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)

            # Create the chart based on the chart_type parameter
            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            elif chart_type == "scatter":
                chart = ScatterChart()
            else:
                raise ValueError(f"Unsupported chart type: {chart_type}")

            # Add the data to the chart
            chart.add_data(data, titles_from_data=True)

            # Set the chart title if provided
            if title:
                chart.title = title

            # Set axis titles if provided
            if x_axis_title:
                chart.x_axis.title = x_axis_title
            if y_axis_title:
                chart.y_axis.title = y_axis_title

            # Handle legend
            if include_legend:
                chart.legend = Legend()
            else:
                chart.legend = None

            # Show data labels if specified
            if show_data_labels:
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showVal = True

            # Add the chart to the sheet at the specified position
            self.sheet.add_chart(chart, position)

            self.logger.info(f"Created {chart_type} chart at position {position}")
        except Exception as e:
            self.logger.error(f"Error creating chart: {str(e)}")
            raise

    def import_from_csv(self, csv_file, start_row=1):
        """
        Import data from a CSV file into the sheet.

        Parameters:
        - csv_file (str): The path to the CSV file.
        - start_row (int): The row number to start importing data from (default is 1).
        """
        with open(csv_file, 'r') as file:
            reader = csv.reader(file)
            for row_num, row in enumerate(reader, start=start_row):
                self.add_row(row, start_row=row_num)

    def import_from_json(self, json_file, start_row=1):
        """
        Import data from a JSON file into the sheet.

        Parameters:
        - json_file (str): The path to the JSON file.
        - start_row (int): The row number to start importing data from (default is 1).
        """
        with open(json_file, 'r') as file:
            data = json.load(file)
            for row_num, row_data in enumerate(data, start=start_row):
                self.add_row(list(row_data.values()), start_row=row_num)

    def protect_sheet(self, password):
        """
        Protect the current sheet with a password.

        Parameters:
        - password (str): The password to protect the sheet.
        """
        self.sheet.protection.password = password

    def add_image(self, image_path, cell):
        """
        Add an image to the sheet at the specified cell.

        Parameters:
        - image_path (str): The path to the image file.
        - cell (str): The cell where the image should be placed (e.g., "A1").
        """
        img = Image(image_path)
        self.sheet.add_image(img, cell)

    def add_hyperlink(self, cell, url, display_text=None):
        """
        Add a hyperlink to a cell.

        Parameters:
        - cell (str): The cell reference to add the hyperlink (e.g., "A1").
        - url (str): The URL for the hyperlink.
        - display_text (str, optional): The text to display for the hyperlink. If None, the URL will be displayed.
        """
        self.sheet[cell].hyperlink = url
        self.sheet[cell].value = display_text if display_text else url

    def apply_zebra_striping(self, start_row, end_row, color1="FFFFFF", color2="F0F0F0"):
        """
        Apply alternating background colors to rows for better readability.

        Parameters:
        - start_row (int): The first row to apply striping.
        - end_row (int): The last row to apply striping.
        - color1 (str): The background color for odd rows (default white).
        - color2 (str): The background color for even rows (default light grey).
        """
        for row in range(start_row, end_row + 1):
            fill_color = PatternFill(start_color=color1, end_color=color1, fill_type="solid") if row % 2 else PatternFill(start_color=color2, end_color=color2, fill_type="solid")
            for cell in self.sheet[row]:
                cell.fill = fill_color

    def apply_cell_style(self, cell_range, font=None, fill=None, border=None, alignment=None):
        """
        Apply custom styles to a cell or range of cells.

        Parameters:
        - cell_range (str): The cell or range to apply styles to (e.g., "A1" or "A1:B10").
        - font (Font, optional): Custom font style.
        - fill (PatternFill, optional): Custom fill style.
        - border (Border, optional): Custom border style.
        - alignment (Alignment, optional): Custom alignment style.
        """
        try:
            for row in self.sheet[cell_range]:
                for cell in row:
                    if font:
                        cell.font = font
                    if fill:
                        cell.fill = fill
                    if border:
                        cell.border = border
                    if alignment:
                        cell.alignment = alignment
            self.logger.info(f"Applied custom style to range: {cell_range}")
        except Exception as e:
            self.logger.error(f"Error applying cell style: {str(e)}")

    def apply_formula(self, cell_range, formula):
        """
        Apply a formula to a cell or range of cells.

        Parameters:
        - cell_range (str): The cell or range to apply the formula to (e.g., "A1" or "A1:A10").
        - formula (str): The formula to apply (e.g., "=SUM(B1:B10)").
        """
        try:
            for row in self.sheet[cell_range]:
                for cell in row:
                    cell.value = formula
            self.logger.info(f"Applied formula '{formula}' to range: {cell_range}")
        except Exception as e:
            self.logger.error(f"Error applying formula: {str(e)}")

    def create_named_range(self, name, cell_range):
        """
        Create a named range in the workbook.

        Parameters:
        - name (str): The name for the range.
        - cell_range (str): The cell range for the named range (e.g., "A1:B10").
        """
        try:
            self.workbook.create_named_range(name, self.sheet, cell_range)
            self.logger.info(f"Created named range '{name}' for range: {cell_range}")
        except Exception as e:
            self.logger.error(f"Error creating named range: {str(e)}")

    def copy_sheet(self, source_sheet_name, new_sheet_name):
        """
        Copy a sheet within the workbook.

        Parameters:
        - source_sheet_name (str): The name of the sheet to copy.
        - new_sheet_name (str): The name for the new sheet.
        """
        try:
            source_sheet = self.workbook[source_sheet_name]
            new_sheet = self.workbook.copy_worksheet(source_sheet)
            new_sheet.title = new_sheet_name
            self.logger.info(f"Copied sheet '{source_sheet_name}' to '{new_sheet_name}'")
        except Exception as e:
            self.logger.error(f"Error copying sheet: {str(e)}")

    def add_cell_comment(self, cell, comment_text, author="ExcelCreator"):
        """
        Add a comment to a cell.

        Parameters:
        - cell (str): The cell to add the comment to (e.g., "A1").
        - comment_text (str): The text of the comment.
        - author (str, optional): The author of the comment.
        """
        try:
            self.sheet[cell].comment = openpyxl.comments.Comment(comment_text, author)
            self.logger.info(f"Added comment to cell {cell}")
        except Exception as e:
            self.logger.error(f"Error adding cell comment: {str(e)}")

    def set_page_setup(self, orientation="portrait", paper_size=9, fit_to_page=False):
        """
        Set up page layout for printing.

        Parameters:
        - orientation (str): Page orientation ("portrait" or "landscape").
        - paper_size (int): Paper size (9 for A4, 1 for Letter, etc.).
        - fit_to_page (bool): Whether to fit the content to one page.
        """
        try:
            self.sheet.page_setup.orientation = orientation
            self.sheet.page_setup.paperSize = paper_size
            if fit_to_page:
                self.sheet.page_setup.fitToPage = True
                self.sheet.page_setup.fitToHeight = 1
                self.sheet.page_setup.fitToWidth = 1
            self.logger.info("Set page setup")
        except Exception as e:
            self.logger.error(f"Error setting page setup: {str(e)}")

    def create_pivot_table(self, source_data, pivot_table_range, rows, cols, values):
        """
        Create a pivot table.

        Parameters:
        - source_data (str): The range containing the source data (e.g., "A1:D100").
        - pivot_table_range (str): The cell where the pivot table should start (e.g., "G1").
        - rows (list): Fields to use for row labels.
        - cols (list): Fields to use for column labels.
        - values (list): Fields to summarize in the pivot table.
        """
        try:
            pivot_sheet = self.workbook.create_sheet("PivotTable")
            pivot_sheet.sheet_view.showGridLines = False

            data_sheet = self.sheet
            pc = openpyxl.pivot.cache.CacheDefinition(cacheSource=f"'{data_sheet.title}'!{source_data}")
            pt = openpyxl.pivot.table.PivotTable(name="PivotTable1", cache=pc)

            for row in rows:
                pt.add_rows(row)
            for col in cols:
                pt.add_columns(col)
            for val in values:
                pt.add_data(val)

            pivot_sheet.add_pivot_table(pt, pivot_table_range)
            self.logger.info(f"Created pivot table in range: {pivot_table_range}")
        except Exception as e:
            self.logger.error(f"Error creating pivot table: {str(e)}")

    def apply_autofilter(self, cell_range):
        """
        Apply AutoFilter to a range of cells.

        Parameters:
        - cell_range (str): The range to apply AutoFilter to (e.g., "A1:D10").
        """
        try:
            self.sheet.auto_filter.ref = cell_range
            self.logger.info(f"Applied AutoFilter to range: {cell_range}")
        except Exception as e:
            self.logger.error(f"Error applying AutoFilter: {str(e)}")

    def save(self):
        """
        Save the Excel workbook to the specified file.
        """
        try:
            self.workbook.save(self.file_name)
            self.logger.info(f"Workbook saved successfully: {self.file_name}")
        except Exception as e:
            self.logger.error(f"Error saving workbook: {str(e)}")
