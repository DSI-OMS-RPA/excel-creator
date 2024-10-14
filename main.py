
from excel_creator import ExcelCreator
from openpyxl.styles import Font, PatternFill

# Example usage
if __name__ == "__main__":

    # Define file name
    file_name = "custom_excel.xlsx"

    # Create an ExcelCreator instance (include headers, not appending to an existing file)
    # excel_creator = ExcelCreator(file_name, include_header=True)

    # Custom header font and background color
    custom_header_font = Font(bold=True, color="000000")  # Black text
    custom_bg_color = "FFFF00"  # Yellow background

    # Create ExcelCreator with custom styles
    excel_creator = ExcelCreator(file_name=file_name, header_font=custom_header_font, header_bg_color=custom_bg_color)

    # Set the sheet name
    excel_creator.set_sheet_name("Payments")

    # Define headers
    headers = [
        "ROW_NUMBER", "PAGAMENTO_ID", "DATA_PAGAMENTO", "CRIADO_POR", "EMPRESA_FACTURACAO",
        "PRODUTO", "ITEM", "COD_MATERIAL", "PRECO_BASE", "IMPOSTO", "DESCONTO",
        "TOTAL", "METODO_PAGAMENTO", "ESTADO", "ENCOMENDA", "PAGAMENTO_AGREGADOR",
        "COD_LOJA_SAP", "NOME_LOJA", "COD_CLIENTE", "NOME_CLIENTE", "CLASSE_PAGAMENTO"
    ]

    # Add headers with starting from row (the internal header font and color will be applied)
    excel_creator.add_headers(headers, start_row=1)

    # Auto-size the columns based on the content
    excel_creator.set_column_widths(auto_size=True)

    # Add a few sample rows of data with a custom font
    data_font = Font(size=11)
    sample_rows = [
        [
            None, 12345, "2024-10-10", "admin", "Empresa XYZ", "Produto A", 1, "M123", 100.0, 10.0, 5.0,
            95.0, "Cartao", "Pago", "ENC001", "AGREG001", "SAP001", "Loja Central", "C001", "Cliente 1", "CLASSE_A"
        ],
        [
            None, 12346, "2024-10-11", "user1", "Empresa ABC", "Produto B", 2, "M124", 200.0, 20.0, 10.0,
            190.0, "Boleto", "Aguardando", "ENC002", "AGREG002", "SAP002", "Loja Norte", "C002", "Cliente 2", "CLASSE_B"
        ]
    ]
    for i, row in enumerate(sample_rows, start=2):
        excel_creator.add_row(row, font=data_font, start_row=i)

    # Apply zebra striping
    excel_creator.apply_zebra_striping(start_row=2, end_row=i)

    # Apply a filter to the data
    excel_creator.apply_cell_style("A1:C1", font=Font(bold=True), fill=PatternFill(start_color="AAAAAA", end_color="AAAAAA", fill_type="solid"))

    # Auto-number the rows in the first column
    excel_creator.auto_number_rows(start_row=2, column=1)

    # Add a comment to a cell
    excel_creator.add_cell_comment("A1", "This is the name column")

    # Apply conditional formatting
    excel_creator.apply_conditional_formatting("K2:K3", min_color="FFCCCC", max_color="00FF00")

    # Apply a formula to the 'DATA_PAGAMENTO' column
    excel_creator.apply_formula("D2:D3", "=B2*2")

    # Add data validation (drop-down list) to the 'METODO_PAGAMENTO' column
    payment_methods = '"Cartao,Dinheiro"'
    excel_creator.add_data_validation("M2:M10", validation_type="list", formula1=payment_methods)

    # Merge cells for a title spanning multiple columns
    # excel_creator.merge_cells("A1:U1")

    # Freeze the first row (headers) and first column
    excel_creator.freeze_panes("B2")

    # Set the page orientation to landscape
    excel_creator.set_page_setup(orientation="landscape")

    # Apply an autofilter
    excel_creator.apply_autofilter("A1:C3")

    # Add a chart (Creating a bar chart with data labels and a custom title)
    excel_creator.create_chart(
        min_col=9, min_row=1, max_col=11, max_row=3,  # Adjusted to match your actual data
        chart_type="bar",
        title="Price, Tax, and Discount Comparison",
        x_axis_title="Category",
        y_axis_title="Amount",
        position="H10",  # Moved to give more space for your data
        include_legend=True,
        show_data_labels=True
    )

    # Save the file
    excel_creator.save()

    print(f"Excel file '{file_name}' created successfully.")
